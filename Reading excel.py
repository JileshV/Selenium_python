import openpyxl

file = "D:\Data.xlsx"
Excel = openpyxl.load_workbook(file)
sheet = Excel["Sheet1"]

rows = sheet.max_row
column = sheet.max_column

for r in range(1, rows+1):
    for c in range(1, column+1):
        cellValue = sheet.cell(r,c).value
        print(cellValue, end="")
        # if type(cellValue) == int:
        cellValue = str(cellValue)
        if len(cellValue) < 15:
            print(" "* (15 - len(cellValue)), end="")
    print("\n")