import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

def readData(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowNum,columnNum).value

def writeData(file, sheetName, data, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNum, columnNum).value = data
    workbook.save(file)

def fillGreen(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheetName = workbook[sheetName]
    greenFill = PatternFill(start_color= "60b212", end_color= "60b212", fill_type= "solid")
    sheetName.cell(rowNum, columnNum).fill = greenFill
    workbook.save(file)

def fillRed(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color= "ff0000", end_color= "ff0000", fill_type= "solid")
    sheet.cell(rowNum, columnNum).fill = redFill
    workbook.save(file)